# -*- coding: utf-8 -*-
"""从项目根目录的「明细台账.xlsx」导出 client 使用的 detailLedgerData.json。"""
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "明细台账.xlsx"
OUT = ROOT / "client" / "src" / "pages" / "overview" / "detailLedgerData.json"

ANNUAL_KEYS = ["budgetYear", "internal", "expectedOccur", "actualOccur", "actualPaid"]
MONTH_KEYS = ["internal", "expectedOccur", "actualOccur", "actualPaid", "unpaid"]


def cell_val(v):
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    return v


def main():
    if not XLSX.exists():
        print(f"未找到: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    r1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    meta = {
        "projectTitle": cell_val(ws.cell(1, 2).value),
        "amountExTaxTotal": cell_val(ws.cell(1, 3).value),
        "annualLabel": cell_val(ws.cell(1, 6).value) or "2023全年",
    }

    months = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, datetime):
            months.append(v.strftime("%Y-%m"))

    rows_out = []
    for r in range(3, ws.max_row + 1):
        subject = ws.cell(r, 2).value
        if subject is None and ws.cell(r, 3).value is None:
            continue
        subj_str = str(subject).strip() if subject is not None else ""
        is_total = "总计" in subj_str or r == ws.max_row

        annual = {
            "budgetYear": cell_val(ws.cell(r, 5).value),
            "internal": cell_val(ws.cell(r, 6).value),
            "expectedOccur": cell_val(ws.cell(r, 7).value),
            "actualOccur": cell_val(ws.cell(r, 8).value),
            "actualPaid": cell_val(ws.cell(r, 9).value),
        }

        month_blocks = []
        for mi, _m in enumerate(months):
            start = 10 + mi * 5
            month_blocks.append(
                {
                    "internal": cell_val(ws.cell(r, start).value),
                    "expectedOccur": cell_val(ws.cell(r, start + 1).value),
                    "actualOccur": cell_val(ws.cell(r, start + 2).value),
                    "actualPaid": cell_val(ws.cell(r, start + 3).value),
                    "unpaid": cell_val(ws.cell(r, start + 4).value),
                }
            )

        rows_out.append(
            {
                "subject": subj_str or None,
                "amountExTax": cell_val(ws.cell(r, 3).value),
                "taxRate": cell_val(ws.cell(r, 4).value),
                "annual": annual,
                "months": month_blocks,
                "isTotal": is_total,
            }
        )

    out = {
        "meta": meta,
        "months": months,
        "annualFieldKeys": ANNUAL_KEYS,
        "monthFieldKeys": MONTH_KEYS,
        "rows": rows_out,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"已写入 {OUT}（{len(months)} 个月 × {len(rows_out)} 行）")


if __name__ == "__main__":
    main()
