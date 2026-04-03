# -*- coding: utf-8 -*-
"""从项目根目录的「成本还原.xlsx」导出 client 使用的 costRestoreData.json。"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "成本还原.xlsx"
OUT = ROOT / "client" / "src" / "pages" / "overview" / "costRestoreData.json"


def cell_val(v):
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("#"):
        return None
    return v


def main():
    if not XLSX.exists():
        print(f"未找到: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    months = []
    for c in range(11, 50, 3):
        d = ws.cell(2, c).value
        if isinstance(d, (int, float)) and d > 40000:
            months.append(from_excel(d).strftime("%Y-%m"))

    rows_out = []
    for r in range(3, 27):
        subject = ws.cell(r, 2).value
        metric = ws.cell(r, 4).value
        yearly = cell_val(ws.cell(r, 6).value)
        occurred = cell_val(ws.cell(r, 8).value)
        unoccurred = cell_val(ws.cell(r, 10).value)
        month_vals = []
        for i in range(len(months)):
            start = 11 + i * 3
            month_vals.append(
                {
                    "paid": cell_val(ws.cell(r, start).value),
                    "expected": cell_val(ws.cell(r, start + 1).value),
                    "unpaid": cell_val(ws.cell(r, start + 2).value),
                }
            )
        rows_out.append(
            {
                "subject": subject,
                "metricType": metric,
                "yearlyTotal": yearly,
                "occurredPaid": occurred,
                "unoccurredUnpaid": unoccurred,
                "months": month_vals,
            }
        )

    out = {"months": months, "rows": rows_out}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"已写入 {OUT}（{len(months)} 个月 × {len(rows_out)} 行）")


if __name__ == "__main__":
    main()
