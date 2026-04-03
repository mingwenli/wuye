# -*- coding: utf-8 -*-
"""从项目根目录的「能耗台账.xlsx」导出 client 使用的 energyLedgerData.json。"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "能耗台账.xlsx"
OUT = ROOT / "client" / "src" / "pages" / "overview" / "energyLedgerData.json"


def cell_val(v):
    if v is None:
        return None
    if isinstance(v, str) and (v.startswith("#") or "DIV/0" in v or "REF!" in v):
        return None
    return v


def main():
    if not XLSX.exists():
        print(f"未找到: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active

    title = cell_val(ws.cell(1, 1).value) or "能耗明细表"
    group_year_mix = cell_val(ws.cell(1, 4).value) or "全年（已发生+预估发生金额）"
    group_rolling = cell_val(ws.cell(1, 7).value) or "全年滚动已发生"
    group_unoccurred = cell_val(ws.cell(1, 10).value) or "全年累计未发生"

    # Row1 month headers: cols 13, 18, 23, ... every 5 cols, 12 months
    month_labels = []
    for c in range(13, 13 + 12 * 5, 5):
        v = ws.cell(1, c).value
        if v is not None:
            month_labels.append(str(v).strip())

    rows_out = []
    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        subj_a = str(a).strip() if a is not None else None
        subj_b = str(b).strip() if b is not None else None
        subj_c = str(c3).strip() if c3 is not None else None

        if subj_a is None and subj_b is None and subj_c is None:
            continue

        annual_year_mix = {
            "budget": cell_val(ws.cell(r, 4).value),
            "manage": cell_val(ws.cell(r, 5).value),
            "completionRate": cell_val(ws.cell(r, 6).value),
        }
        annual_rolling = {
            "budget": cell_val(ws.cell(r, 7).value),
            "actualOccur": cell_val(ws.cell(r, 8).value),
            "remainDiff": cell_val(ws.cell(r, 9).value),
        }
        annual_unoccurred = {
            "budget": cell_val(ws.cell(r, 10).value),
            "expectedOccur": cell_val(ws.cell(r, 11).value),
            "remainDiff": cell_val(ws.cell(r, 12).value),
        }

        month_blocks = []
        for mi in range(len(month_labels)):
            start = 13 + mi * 5
            month_blocks.append(
                {
                    "budget": cell_val(ws.cell(r, start).value),
                    "expectedOccur": cell_val(ws.cell(r, start + 1).value),
                    "actualOccur": cell_val(ws.cell(r, start + 2).value),
                    "completionRate": cell_val(ws.cell(r, start + 3).value),
                    "actualPaid": cell_val(ws.cell(r, start + 4).value),
                }
            )

        is_summary = r == 3

        rows_out.append(
            {
                "metricName": subj_a,
                "energyType": subj_b,
                "metricKind": subj_c,
                "annualYearMix": annual_year_mix,
                "annualRolling": annual_rolling,
                "annualUnoccurred": annual_unoccurred,
                "months": month_blocks,
                "isSummaryRow": bool(is_summary),
            }
        )

    out = {
        "meta": {
            "title": title,
            "monthLabels": month_labels,
            "groupYearMix": group_year_mix,
            "groupRolling": group_rolling,
            "groupUnoccurred": group_unoccurred,
        },
        "rows": rows_out,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"已写入 {OUT}（{len(month_labels)} 个月 × {len(rows_out)} 行）")


if __name__ == "__main__":
    main()
